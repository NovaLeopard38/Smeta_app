import io
import pytest
from openpyxl import load_workbook


def test_export_empty_smeta_xlsx(client, auth_headers):
    """Export empty smeta as Excel."""
    smeta = client.post("/smetas", json={"name": "Export Test"}, headers=auth_headers).json()
    smeta_id = smeta["id"]
    token = auth_headers["Authorization"].split(" ")[1]
    resp = client.get(f"/smetas/{smeta_id}/export.xlsx?token={token}")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = load_workbook(io.BytesIO(resp.content))
    assert len(wb.sheetnames) >= 1


def test_export_smeta_with_items_xlsx(client, auth_headers):
    """Export smeta with items and verify content."""
    smeta = client.post("/smetas", json={"name": "С позициями"}, headers=auth_headers).json()
    smeta_id = smeta["id"]
    client.post(f"/smetas/{smeta_id}/items", json={
        "name": "Камера IP",
        "section": "Оборудование",
        "unit": "шт",
        "quantity": 2,
        "unit_price": 5000,
    }, headers=auth_headers)
    token = auth_headers["Authorization"].split(" ")[1]
    resp = client.get(f"/smetas/{smeta_id}/export.xlsx?token={token}")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert any("Камера" in str(v) for v in values if v)


def test_print_smeta_html(client, auth_headers):
    """Print smeta returns HTML with smeta name."""
    smeta = client.post("/smetas", json={"name": "Print Test"}, headers=auth_headers).json()
    token = auth_headers["Authorization"].split(" ")[1]
    resp = client.get(f"/smetas/{smeta['id']}/print?token={token}")
    assert resp.status_code == 200
    assert "text/html" in resp.headers["content-type"]
    assert "Print Test" in resp.text
