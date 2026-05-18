
import json
import re
from html import escape
from io import BytesIO

import httpx
import pandas as pd
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .text_utils import (
    clean_text,
    classify_catalog_item,
    endpoint,
    extract_price,
    extract_strict_price,
    is_numeric_price,
    normalize_label,
    normalize_quantity,
    summarize_characteristics,
)
from config import DEFAULT_SECTIONS, read_settings
from crud import create_material


def parse_section_adjustments(value):
    if isinstance(value, dict):
        raw = value
    else:
        try:
            raw = json.loads(value or "{}")
        except (TypeError, json.JSONDecodeError):
            raw = {}
    adjustments = {}
    for section, percent in raw.items():
        try:
            adjustments[str(section)] = max(-100.0, min(1000.0, float(percent or 0)))
        except (TypeError, ValueError):
            adjustments[str(section)] = 0
    return adjustments


def section_adjustment_percent(smeta, section):
    return parse_section_adjustments(getattr(smeta, "section_adjustments", "{}")).get(section or "Оборудование", 0)


def effective_unit_price(item, smeta):
    percent = section_adjustment_percent(smeta, item.section or "Оборудование")
    return round((item.unit_price or 0) * (1 + percent / 100), 2)


def item_total(item, smeta):
    return round((item.quantity or 0) * effective_unit_price(item, smeta), 2)


def smeta_financials(smeta):
    subtotal = round(sum(item_total(item, smeta) for item in smeta.items), 2)
    tax_mode = getattr(smeta, "tax_mode", "none") or "none"
    tax_rate = float(getattr(smeta, "tax_rate", 0) or 0)
    if tax_mode == "vat_added" and tax_rate > 0:
        tax_amount = round(subtotal * tax_rate / 100, 2)
        total = round(subtotal + tax_amount, 2)
    elif tax_mode == "vat_included" and tax_rate > 0:
        total = subtotal
        tax_amount = round(subtotal * tax_rate / (100 + tax_rate), 2)
    else:
        tax_amount = 0
        total = subtotal
    return {"subtotal": subtotal, "tax_amount": tax_amount, "total": total}


def normalized_parent_id(smeta):
    try:
        return int(getattr(smeta, "parent_id", None) or 0) or None
    except (TypeError, ValueError):
        return None


def item_to_dict(item, smeta=None):
    price = effective_unit_price(item, smeta) if smeta else (item.unit_price or 0)
    base_value = getattr(item, "base_unit_price", None)
    base_price = float(base_value if base_value is not None else (item.unit_price or 0))
    return {
        "id": item.id,
        "item_type": item.item_type,
        "section": item.section or "Оборудование",
        "name": item.name,
        "characteristics": item.characteristics or "",
        "unit": item.unit or "",
        "quantity": item.quantity,
        "unit_price": item.unit_price,
        "base_unit_price": base_price,
        "effective_unit_price": price,
        "section_adjustment_percent": section_adjustment_percent(smeta, item.section or "Оборудование") if smeta else 0,
        "source": item.source or "",
        "total": round((item.quantity or 0) * price, 2),
    }


def smeta_to_dict(smeta):
    items = [item_to_dict(item, smeta) for item in smeta.items]
    financials = smeta_financials(smeta)
    return {
        "id": smeta.id,
        "parent_id": normalized_parent_id(smeta),
        "owner_id": _normalized_owner_id(smeta),
        "is_branch": bool(normalized_parent_id(smeta)),
        "name": smeta.name,
        "customer_name": smeta.customer_name or "",
        "customer_details": smeta.customer_details or "",
        "contractor_name": smeta.contractor_name or "",
        "contractor_details": smeta.contractor_details or "",
        "approver_name": smeta.approver_name or "",
        "approver_details": smeta.approver_details or "",
        "tax_mode": getattr(smeta, "tax_mode", "none") or "none",
        "tax_rate": float(getattr(smeta, "tax_rate", 0) or 0),
        "section_adjustments": parse_section_adjustments(getattr(smeta, "section_adjustments", "{}")),
        "created_at": smeta.created_at,
        "items": items,
        "subtotal": financials["subtotal"],
        "tax_amount": financials["tax_amount"],
        "total": financials["total"],
    }


def _normalized_owner_id(smeta):
    try:
        return int(getattr(smeta, "owner_id", None) or 0) or None
    except (TypeError, ValueError):
        return None


def first_matching_column(columns, variants):
    normalized = [(column, str(column).lower()) for column in columns]
    for column, lower_column in normalized:
        if any(variant in lower_column for variant in variants):
            return column
    return None


def dataframe_to_text(df):
    df = df.dropna(how="all").fillna("")
    return df.head(160).to_csv(index=False)


def header_score(row):
    labels = [normalize_label(value) for value in row if clean_text(value)]
    text_row = " | ".join(labels)
    score = 0
    if any(word in text_row for word in ["наименование", "название", "номенклатура", "товар", "работ"]):
        score += 3
    if any(word in text_row for word in ["цена", "цены", "розн", "опт", "парт", "стоим"]):
        score += 3
    if any(word in text_row for word in ["ед.", "ед ", "изм", "руб./", "вал."]):
        score += 1
    if any(word in text_row for word in ["описание", "характерист", "краткие"]):
        score += 1
    return score


def find_header_row(rows):
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:80]):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index if best_score >= 4 else None


def column_values(rows, column_index, start_index, limit=80):
    values = []
    for row in rows[start_index : start_index + limit]:
        if column_index < len(row):
            values.append(row[column_index])
    return values


def detect_name_column(headers, rows, data_start):
    candidates = []
    for index, header in enumerate(headers):
        label = normalize_label(header)
        score = 0
        if any(word in label for word in ["наименование", "название", "номенклатура", "товар"]):
            score += 100
        if "работ" in label:
            score += 80
        values = column_values(rows, index, data_start)
        text_count = sum(1 for value in values if len(clean_text(value)) >= 3 and not is_numeric_price(value))
        score += text_count
        if score:
            candidates.append((score, index))
    return max(candidates)[1] if candidates else None


def detect_characteristics_column(headers):
    for index, header in enumerate(headers):
        label = normalize_label(header)
        if any(word in label for word in ["характер", "описан", "кратк", "параметр", "модель"]):
            return index
    return None


def detect_unit_column(headers):
    for index, header in enumerate(headers):
        label = normalize_label(header)
        if any(word in label for word in ["ед.", "ед ", "изм", "вал./", "руб./"]):
            return index
    return None


def price_header_priority(header):
    label = normalize_label(header)
    if "парт" in label:
        return 100
    if "кр.опт" in label or "круп" in label:
        return 95
    if "опт" in label:
        return 90
    if "цена" in label or "цены" in label or "стоим" in label:
        return 75
    if "инст" in label:
        return 55
    if "розн" in label or "ритейл" in label or "rrp" in label or "rrc" in label or "rrc" in label:
        return 110
    return 0


def detect_price_columns(headers, parent_headers, rows, data_start):
    candidates = []
    for index, header in enumerate(headers):
        combined_header = f"{clean_text(parent_headers[index])} {clean_text(header)}"
        header_priority = price_header_priority(combined_header)
        values = column_values(rows, index, data_start, limit=120)
        numeric_count = sum(1 for value in values if is_numeric_price(value))
        if header_priority or numeric_count >= 3:
            if any(word in normalize_label(combined_header) for word in ["код", "артикул", "№", "номер"]):
                continue
            score = header_priority + numeric_count
            candidates.append((score, index))
    return [index for _, index in sorted(candidates, reverse=True)]


def likely_category_row(row, name_index, price_indexes):
    name = clean_text(row[name_index] if name_index is not None and name_index < len(row) else "")
    if not name:
        return False
    has_price = any(index < len(row) and is_numeric_price(row[index]) for index in price_indexes)
    filled = sum(1 for value in row if clean_text(value))
    return not has_price and filled <= 3 and len(name) >= 3


def pick_price(row, price_indexes):
    # VSB: user wants RETAIL (highest) price by default — not wholesale.
    prices = []
    for index in price_indexes:
        if index < len(row):
            price = extract_strict_price(row[index])
            if price is not None and price > 0:
                prices.append(price)
    if not prices:
        return None
    return max(prices)


def parse_excel_workbook(file_obj, source):
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    parsed = []
    for sheet in workbook.worksheets:
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        header_index = find_header_row(rows)
        if header_index is None:
            continue

        parent_headers = rows[header_index]
        next_headers = rows[header_index + 1] if header_index + 1 < len(rows) else ()
        use_next_for_price = any(normalize_label(value) == "цены" for value in parent_headers)
        headers = []
        for index, header in enumerate(parent_headers):
            next_header = next_headers[index] if index < len(next_headers) else ""
            headers.append(f"{clean_text(header)} {clean_text(next_header) if use_next_for_price else ''}".strip())

        data_start = header_index + (2 if use_next_for_price else 1)
        name_index = detect_name_column(headers, rows, data_start)
        if name_index is None:
            continue
        characteristics_index = detect_characteristics_column(headers)
        unit_index = detect_unit_column(headers)
        price_indexes = detect_price_columns(headers, parent_headers, rows, data_start)
        if not price_indexes:
            continue

        current_group = sheet.title
        for row in rows[data_start:]:
            row = tuple(row)
            if likely_category_row(row, name_index, price_indexes):
                current_group = clean_text(row[name_index])
                continue

            name = clean_text(row[name_index] if name_index < len(row) else "")
            price = pick_price(row, price_indexes)
            if not name or price is None:
                continue
            if name.lower() in {"наименование", "название", "товар", "цена"}:
                continue

            characteristics = (
                clean_text(row[characteristics_index])
                if characteristics_index is not None and characteristics_index < len(row)
                else ""
            )
            unit = clean_text(row[unit_index]) if unit_index is not None and unit_index < len(row) else ""
            if unit.lower().startswith("руб./"):
                unit = unit.split("/", 1)[-1]
            parsed.append(
                {
                    "name": name,
                    "characteristics": characteristics or current_group,
                    "unit": unit,
                    "price": price,
                    "source": f"{source} / {sheet.title}",
                    "item_type": classify_catalog_item(name, f"{source} / {sheet.title}"),
                }
            )
    return parsed


def save_parsed_materials(db, rows):
    imported = 0
    skipped = 0
    seen = set()
    for row in rows:
        key = (row["name"], row.get("source", ""), row.get("price"))
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        create_material(
            db,
            row["name"],
            row.get("unit", ""),
            float(row["price"]),
            row.get("source", ""),
            row.get("characteristics", ""),
            row.get("item_type", "equipment"),
        )
        imported += 1
    return imported, skipped


def import_excel_by_guess(db, df, source):
    name_column = first_matching_column(
        df.columns,
        ["name", "назван", "наимен", "номенклат", "товар", "материал", "позиция"],
    )
    price_column = first_matching_column(
        df.columns,
        ["price", "цена", "стоим", "прайс", "розн", "опт"],
    )
    unit_column = first_matching_column(df.columns, ["unit", "ед", "изм"])
    characteristics_column = first_matching_column(
        df.columns,
        ["характер", "описан", "параметр", "модель", "артикул"],
    )

    if not name_column or not price_column:
        return 0, 0

    imported = 0
    skipped = 0
    for _, row in df.dropna(how="all").iterrows():
        name = "" if pd.isna(row.get(name_column)) else str(row.get(name_column)).strip()
        price = extract_price(row.get(price_column))
        if not name or price is None:
            skipped += 1
            continue
        unit = "" if not unit_column or pd.isna(row.get(unit_column)) else str(row.get(unit_column)).strip()
        characteristics = (
            ""
            if not characteristics_column or pd.isna(row.get(characteristics_column))
            else str(row.get(characteristics_column)).strip()
        )
        create_material(db, name, unit, price, source, characteristics, classify_catalog_item(name, source))
        imported += 1
    return imported, skipped


async def extract_pdf_text(file):
    settings = read_settings()
    if not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="Для PDF нужен API-ключ AI-провайдера")
    content = await file.read()
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            response = await client.post(
                endpoint(settings["base_url"], "extract_text"),
                headers={"Authorization": f"Bearer {settings['api_key']}"},
                files={"file": (file.filename, content, file.content_type or "application/pdf")},
            )
            response.raise_for_status()
    except httpx.HTTPError as exc:
        from .text_utils import http_error_detail
        raise HTTPException(status_code=502, detail=http_error_detail(exc, "Не удалось извлечь текст из PDF")) from exc
    data = response.json()
    if isinstance(data, dict):
        return data.get("text") or data.get("content") or data.get("result") or json.dumps(data, ensure_ascii=False)
    return str(data)


def save_ai_materials(db, rows, source):
    imported = 0
    skipped = 0
    for row in rows:
        if not isinstance(row, dict):
            skipped += 1
            continue
        name = str(row.get("name") or row.get("название") or "").strip()
        price = extract_price(row.get("price") or row.get("цена"))
        if not name or price is None:
            skipped += 1
            continue
        create_material(
            db,
            name,
            str(row.get("unit") or row.get("единица") or "").strip(),
            price,
            str(row.get("source") or source or "").strip(),
            str(row.get("characteristics") or row.get("характеристики") or "").strip(),
            classify_catalog_item(name, source),
        )
        imported += 1
    return imported, skipped


def smeta_check_issues(smeta):
    issues = []
    if not smeta:
        return issues
    for item in smeta.items:
        if (item.unit_price or 0) == 0:
            issues.append(f"Нулевая цена: {item.section or 'Раздел'} / {item.name}")
        if (item.quantity or 0) <= 0:
            issues.append(f"Некорректное количество: {item.name}")
    return issues


def grouped_smeta_items(smeta):
    groups = []
    for section in DEFAULT_SECTIONS:
        items = [item for item in smeta.items if (item.section or "Оборудование") == section]
        if items:
            groups.append((section, items))
    extra_items = [
        item
        for item in smeta.items
        if (item.section or "Оборудование") not in DEFAULT_SECTIONS
    ]
    if extra_items:
        groups.append(("Прочее", extra_items))
    return groups


def org_rows(smeta):
    return [
        ("Заказчик", smeta.customer_name or "", smeta.customer_details or ""),
        ("Исполнитель", smeta.contractor_name or "", smeta.contractor_details or ""),
        ("Согласующий", smeta.approver_name or "", smeta.approver_details or ""),
    ]


def safe_filename(value):
    value = re.sub(r"[^0-9a-zA-Zа-яА-ЯёЁ._ -]+", "_", value or "smeta").strip()
    return value[:80] or "smeta"


def build_smeta_workbook(smeta):
    financials = smeta_financials(smeta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.sheet_view.showGridLines = False
    details_ws = wb.create_sheet("Характеристики")
    details_ws.sheet_view.showGridLines = False

    widths = [5, 42, 56, 8, 9, 12, 13]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    def solid(color):
        return PatternFill(fill_type="solid", fgColor=f"FF{color}")

    title_fill = solid("173B57")
    section_fills = {
        "Оборудование": solid("B6D7A8"),
        "Монтажные работы": solid("F6D776"),
        "Пусконаладочные работы": solid("D9C2E9"),
        "Кабельные линии": solid("C9E2B3"),
        "Материалы и расходники": solid("F4C7B6"),
        "Доставка и логистика": solid("BDD7EE"),
        "Проектирование": solid("D9D9D9"),
        "Прочее": solid("E2E2E2"),
    }
    item_fills = {
        "Оборудование": solid("D9EAD3"),
        "Монтажные работы": solid("FFF2CC"),
        "Пусконаладочные работы": solid("EADCF8"),
        "Кабельные линии": solid("E2F0D9"),
        "Материалы и расходники": solid("FCE4D6"),
        "Доставка и логистика": solid("DDEBF7"),
        "Проектирование": solid("EDEDED"),
        "Прочее": solid("F2F2F2"),
    }
    section_fill = solid("DDEBF7")
    header_fill = solid("D9EAF7")
    org_fill = solid("EEF3F7")
    org_role_fill = solid("DDEAF6")
    total_fill = solid("1F4E78")
    total_light_fill = solid("E8F3FF")
    signature_fill = solid("FAFAFA")
    thin = Side(style="thin", color="9FB2C3")
    medium = Side(style="medium", color="6E879F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=medium, right=thin, top=medium, bottom=thin)
    money_format = '#,##0.00 "\u20bd"'

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    ws["A1"] = f"Смета: {smeta.name}"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["F1"] = "Итого"
    ws["G1"] = financials["total"]
    for col in range(1, 8):
        cell = ws.cell(1, col)
        cell.fill = title_fill if col <= 5 else total_light_fill
        cell.border = border
        if col <= 5:
            cell.font = Font(bold=True, size=16, color="FFFFFF")
        else:
            cell.font = Font(bold=True, size=13 if col == 6 else 15, color="173B57")
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws["G1"].number_format = money_format
    ws.row_dimensions[1].height = 28

    row = 2
    for col, value in enumerate(["Роль", "Организация", "Реквизиты", "", "", "", ""], 1):
        ws.cell(row, col, value)
    for col in range(1, 8):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, color="173B57")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    for role, name, details in org_rows(smeta):
        ws.cell(row, 1, role)
        ws.cell(row, 2, name or "Не заполнено")
        ws.cell(row, 3, details or "")
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.fill = org_role_fill if col == 1 else org_fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 1).font = Font(bold=True)
        ws.row_dimensions[row].height = 20 if not details else 34
        row += 1

    row += 1
    table_header_row = row
    headers = ["№", "Позиция", "Характеристики", "Ед.", "Кол-во", "Цена", "Сумма"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True, color="173B57")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24
    row += 1

    detail_rows = [["№", "Раздел", "Позиция", "Полные характеристики", "Источник"]]
    for section_index, (section, items) in enumerate(grouped_smeta_items(smeta)):
        if section_index > 0:
            row += 2
        section_total = sum(item_total(item, smeta) for item in items)
        ws.cell(row, 1, section)
        ws.cell(row, 7, section_total)
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = section_fills.get(section, section_fill)
            cell.border = section_border if col == 1 else border
            cell.alignment = Alignment(vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1
        section_item_no = 1
        for item in items:
            ws.cell(row, 1, section_item_no)
            ws.cell(row, 2, item.name)
            short_characteristics = clean_text(
                summarize_characteristics(item.characteristics or "", max_lines=1, line_length=135)
            )
            ws.cell(row, 3, short_characteristics)
            ws.cell(row, 4, item.unit or "ед.")
            ws.cell(row, 5, item.quantity or 0)
            ws.cell(row, 6, effective_unit_price(item, smeta))
            ws.cell(row, 7, f"=E{row}*F{row}")
            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.fill = item_fills.get(section, solid("FFFFFF"))
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="top")
            ws.row_dimensions[row].height = 42 if short_characteristics else 30
            ws.cell(row, 5).number_format = "0"
            ws.cell(row, 6).number_format = money_format
            ws.cell(row, 7).number_format = money_format
            detail_rows.append(
                [section_item_no, item.section or "", item.name, item.characteristics or "", item.source or ""]
            )
            section_item_no += 1
            row += 1

    last_item_row = row - 1
    row += 1
    if financials["tax_amount"] > 0 and (smeta.tax_mode or "") == "vat_added":
        ws.cell(row, 1, "Итого без НДС")
        ws.cell(row, 7, financials["subtotal"])
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="173B57")
            cell.fill = total_light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1
        ws.cell(row, 1, f"НДС {float(smeta.tax_rate or 0):g}%")
        ws.cell(row, 7, financials["tax_amount"])
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="173B57")
            cell.fill = total_light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1
    elif financials["tax_amount"] > 0 and (smeta.tax_mode or "") == "vat_included":
        ws.cell(row, 1, f"В том числе НДС {float(smeta.tax_rate or 0):g}%")
        ws.cell(row, 7, financials["tax_amount"])
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="173B57")
            cell.fill = total_light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
        ws.cell(row, 7).number_format = money_format
        row += 1

    ws.cell(row, 1, "Полная сумма сметы")
    ws.cell(row, 7, financials["total"])
    for col in range(1, 8):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="right" if col == 6 else "center", vertical="center")
    ws.cell(row, 7).number_format = money_format
    ws.cell(row, 7).font = Font(bold=True, size=16, color="FFFFFF")
    ws.row_dimensions[row].height = 32

    row += 3
    ws.cell(row, 1, "Заказчик").font = Font(bold=True)
    ws.cell(row, 5, "Исполнитель").font = Font(bold=True)
    for col in range(1, 8):
        ws.cell(row, col).fill = signature_fill
    row += 2
    ws.cell(row, 1, "________________ /")
    ws.cell(row, 5, "________________ /")
    for col in range(1, 8):
        ws.cell(row, col).fill = signature_fill

    ws.freeze_panes = f"A{table_header_row + 1}"
    ws.auto_filter.ref = f"A{table_header_row}:G{last_item_row}"

    for index, values in enumerate(detail_rows, 1):
        for col, value in enumerate(values, 1):
            details_ws.cell(index, col, value)
            details_ws.cell(index, col).alignment = Alignment(wrap_text=True, vertical="top")
            details_ws.cell(index, col).border = border
        if index == 1:
            for col in range(1, 6):
                details_ws.cell(index, col).font = Font(bold=True, color="173B57")
                details_ws.cell(index, col).fill = header_fill
    for index, width in enumerate([6, 20, 42, 90, 36], 1):
        details_ws.column_dimensions[get_column_letter(index)].width = width
    details_ws.freeze_panes = "A2"

    return wb
